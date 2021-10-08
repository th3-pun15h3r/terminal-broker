from typing import Text
from zipfile import ZipFile
import os
import pandas as pd
import docx
from openpyxl import load_workbook
import json
import boto3
import fitz
import io
from PIL import Image
import os
from pathlib import Path
from .models import Developer, Locations, Properties, paymentPlan, priceRange, projectAttachment, projectDetails, projectHighlights, propertyImages
# Build paths inside the project like this: BASE_DIR / 'subdir'.
BASE_DIR = Path(__file__).resolve().parent.parent


def Extract(file):
    with ZipFile(file, 'r') as zip:
        # zip.printdir()
        zip.extractall(r"{}/Extract_files".format(BASE_DIR))
        print("Process Completed")
        path = r"{}/Extract_files".format(BASE_DIR)
        for file in os.listdir(path):
            d = os.path.join(path, file)
            if os.path.isdir(d):
                print(file)
                devel = Developer(
                    name=file, logo=""
                )
                devel.save()
                print(devel, "developer")
                for locations in os.listdir(d):
                    l = os.path.join(d, locations)
                    if os.path.isdir(l):
                        print(locations)
                        locations = Locations.objects.create(
                            developer=devel,
                            name=locations,
                        )
                        locations.save()

                        for properti in os.listdir(l):
                            po = os.path.join(l, properti)
                            if os.path.isdir(po):
                                print(properti)
                                property = Properties(
                                    developer=devel,
                                    location=locations,
                                    name=properti,
                                )
                                property.save()
                                for root, dirs, files in os.walk(po):

                                    if root.endswith("/images"):
                                        imageList = extractimages(root)
                                        pImages = propertyImages.objects.create(
                                            property=property, images=imageList)
                                        pImages.save()

                                    for f in files:
                                        # print(os.path.join(root,f))
                                        roots = os.path.join(root, f)

                                        if (f == "project_details.xlsx"):
                                            p_details = project_details(
                                                roots)
                                            p_det = projectDetails.objects.create(
                                                property=property, projectdetails=p_details)
                                            p_det.save()
                                        if (f == "payment_plan.xlsx"):
                                            p_plan = payment_plan(roots)
                                            p_plans = paymentPlan.objects.create(
                                                property=property, paymentplan=p_plan)
                                            p_plans.save()
                                        if f.endswith(".xlsx") and f != "project_details.xlsx" and f != "payment_plan.xlsx":
                                            price = price_range(roots)
                                            price_Det = priceRange.objects.create(
                                                property=property, pricerange=price)
                                            price_Det.save()

                                        if f.endswith(".docx"):
                                            Texts = []
                                            roots = os.path.join(root, f)
                                            doc = docx.Document(roots)
                                            for para in doc.paragraphs:
                                                Texts.append(para.text)
                                            proj_high = projectHighlights.objects.create(
                                                property=property, projecthighlights=Texts)
                                            proj_high.save()
                                        '''        
                                        if f.endswith(".pdf"):
                                            roots = os.path.join(root, f)
                                            pdf_img = pdf(roots, f)
                                            attach = projectAttachment.objects.create(
                                                property=property, attachment=pdf_img)
                                            attach.save()'''

        return True


def extractimages(root):
    imageList = []
    list = os.listdir(root)
    for values in list:
        roots = os.path.join(root, values)
        s3 = boto3.resource(
            service_name='s3',
            region_name='us-east-2',
            aws_access_key_id='AKIAW6UDUMDA4ALRZAHZ',
            aws_secret_access_key='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'
        )
        s3.Bucket('aajproperty').upload_file(Filename=roots,
                                             Key=values)
        imageList.append(
            "https://aajproperty.s3.ap-southeast-1.amazonaws.com/" + values)
    print(imageList, "imagelist")
    imgDIct = {}
    for ind, each in enumerate(imageList):
        imgDIct[each] = str(ind)
    return imgDIct


def project_details(roots):
    json_data = None
    wb = load_workbook(roots)
    ws = wb.active
    name1 = ws.cell(row=1, column=1).value
    value1 = ws.cell(row=2, column=1).value
    name2 = ws.cell(row=1, column=2).value
    value2 = ws.cell(row=2, column=2).value
    name3 = ws.cell(row=1, column=3).value
    value3 = ws.cell(row=2, column=3).value
    name4 = ws.cell(row=1, column=4).value
    value4 = ws.cell(row=2, column=4).value
    name5 = ws.cell(row=1, column=5).value
    value5 = ws.cell(row=2, column=5).value
    name6 = ws.cell(row=1, column=6).value
    value6 = ws.cell(row=2, column=6).value
    name7 = ws.cell(row=1, column=7).value
    value7 = ws.cell(row=2, column=7).value
    name8 = ws.cell(row=1, column=8).value
    value8 = ws.cell(row=2, column=8).value
    name9 = ws.cell(row=1, column=9).value
    value9 = ws.cell(row=2, column=9).value
    data = {
        "Starting_Price": value1,
        "Price_Per_Sqft_from": value2,
        name3: value3,
        name4: value4,
        "Est_Completion": value5,
        name6: value6,
        name7: value7,
        name8: value8,
        name9: value9,
    }
    json_data = json.dumps(data)
    return json_data


def payment_plan(roots):
    payement_list = []
    wb = load_workbook(roots)
    ws = wb.active
    sheet = wb.worksheets[0]
    rows = sheet.max_row
    for i in range(4, rows + 1):
        name = ws.cell(row=i, column=1).value
        milestone = ws.cell(row=i, column=2).value
        payment = ws.cell(row=i, column=3).value

        data = {
            "Installment": name,
            "Milestone": milestone,
            "Payment": payment,
        }
        json_data = json.dumps(data)
        payement_list.append(json_data)
    return payement_list


def price_range(roots):
    plists = []
    wb = load_workbook(roots)
    ws = wb.active
    sheet = wb.worksheets[0]
    rows = sheet.max_row
    for i in range(3, rows, 3):
        name1 = ws.cell(row=i, column=1).value
        value = ws.cell(row=i + 1, column=1).value
        size_name = ws.cell(row=i + 1, column=2).value
        size_value = ws.cell(row=i + 2, column=2).value
        price_name = ws.cell(row=i + 1, column=3).value
        price_value = ws.cell(row=i + 2, column=3).value

        data = {
            "name": name1,
            "BR": value,
            "Size_from_to_(Sqft.)": size_value,
            "Price_from_to_(AED)": price_value,
        }
        json_datas = json.dumps(data)
        print(json_datas, "jsondatas")
        plists.append(json_datas)
    print(plists, "pricerange")
    return plists

'''
def pdf(roots, f):
    pdfimg = []
    pdf_file = fitz.open(roots)
    for page_index in range(len(pdf_file)):
        page = pdf_file[page_index]
        image_list = page.getImageList()
        if image_list:
            print(
                f"[+] Found a total of {len(image_list)} images in page {page_index}")
        for image_index, img in enumerate(page.getImageList(), start=1):
            xref = img[0]
            base_image = pdf_file.extractImage(
                xref)
            image_bytes = base_image["image"]
            image_ext = base_image["ext"]
            image = Image.open(
                io.BytesIO(image_bytes))
            s3 = boto3.resource(
                service_name='s3',
                region_name='us-east-2',
                aws_access_key_id='AKIAW6UDUMDA4ALRZAHZ',
                aws_secret_access_key='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'
            )
            image.save(
                open(f"image{f}{page_index + 1}_{image_index}.{image_ext}", "wb"))
            s3.Bucket('aajproperty').upload_file(
                Filename=f"{BASE_DIR}/image{f}{page_index + 1}_{image_index}.{image_ext}",
                Key=f"image{f}{page_index + 1}_{image_index}.{image_ext}")
            # s3.upload_fileobj(
            #     image, "aajproperty", f"image{f}{page_index + 1}_{image_index}.{image_ext}")
            pdfimg.append("https://aajproperty.s3.ap-southeast-1.amazonaws.com/" +
                          f"image{f}{page_index + 1}_{image_index}.{image_ext}")
            os.remove(
                f"{BASE_DIR}/image{f}{page_index + 1}_{image_index}.{image_ext}")
    print(pdfimg, "pdfimages")
    return pdfimg
'''